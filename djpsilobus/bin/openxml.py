# -*- coding: utf-8 -*-
import os, sys

# env
sys.path.append('/usr/lib/python2.7/dist-packages/')
sys.path.append('/usr/lib/python2.7/')
sys.path.append('/usr/local/lib/python2.7/dist-packages/')
sys.path.append('/data2/django_1.9/')
sys.path.append('/data2/django_projects/')
sys.path.append('/data2/django_third/')
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "djpsilobus.settings")

import django

django.setup()

from django.conf import settings

from djpsilobus.core.utils import syllabus_name
from djpsilobus.core.data import HEADERS
from djzbar.utils.academics import sections

from openpyxl import Workbook
from openpyxl import load_workbook

import argparse

YEAR = settings.YEAR
SESS = settings.SESS

# set up command-line options
desc = """
    Creates an OpenXML workbook with course syllabi data.
    Required: Department code. e.g. PHY
"""

parser = argparse.ArgumentParser(description=desc)

parser.add_argument(
    "--division",
    required=True,
    help="Department code e.q.'NSSS'",
    dest="division"
)
parser.add_argument(
    "--department",
    required=False,
    help="Department code e.q.'PHY'",
    dest="department"
)
parser.add_argument(
    "--debug",
    action='store_true',
    help="Dry run?",
    dest="debug"
)

def main():

    if debug:
        print "department code = {}".format(department)

    courses = sections(code=department,year=YEAR,sess=SESS)

    if courses:
        #wb = Workbook()
        wb = load_workbook('assets/template.xlsx')
        # grab the active worksheet
        ws = wb.active

        #print ws['A1'].value
        # Rename sheet
        ws.title = department
        # 2.4 required for this to work
        new_sheet = wb.copy_worksheet(ws)
        new_sheet.title = "PHY"
        # create a new sheet
        #ws2 = wb.create_sheet(title="PHY")
        #ws2.append(HEADERS)
        # create a list for each row and insert into workbook
        for c in courses:
            section = []
            for course in c:
                section.append(course)

            # check for syllabus
            phile = syllabus_name(c)
            path = "{}{}/{}/{}/{}/{}.pdf".format(
                settings.UPLOADS_DIR,YEAR,SESS,division,department,phile
            )
            if debug:
                print path
            if os.path.isfile(path):
                syllabus="Yes"
            else:
                syllabus="No"

            section.append(syllabus)
            ws.append(section)

        # Save the file
        wb.save("{}.xlsx".format(department))
    else:
        print "no courses found"


######################
# shell command line
######################

if __name__ == "__main__":
    args = parser.parse_args()
    department = args.department
    division = args.division
    debug = args.debug

    sys.exit(main())
